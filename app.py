import streamlit as st
import openpyxl
import pandas as pd
import os
import io
import tempfile
from openpyxl.utils import get_column_letter
import base64

# Configuración de la página
st.set_page_config(
    page_title="Generador de Mapas SVG",
    page_icon="🗺️",
    layout="wide",
    initial_sidebar_state="expanded",
)


def extraer_ubicacion_y_capacidad(valor):
    """
    Extrae ubicación y capacidad de un valor en formato 'UBICACION (CAPACIDAD)'
    """
    if valor is None:
        return None, None

    if isinstance(valor, str) and "(" in valor and ")" in valor:
        try:
            ubicacion = valor.split("(")[0].strip()
            capacidad_str = valor.split("(")[1].split(")")[0].strip()

            try:
                capacidad = int(capacidad_str)
            except ValueError:
                try:
                    capacidad = float(capacidad_str)
                except ValueError:
                    capacidad = None
        except:
            ubicacion = valor
            capacidad = None
    else:
        ubicacion = valor
        capacidad = None

    return ubicacion, capacidad


def generar_layout_general(uploaded_file, ubicaciones_referencia, hoja_origen):
    """
    Genera layout desde Excel con procesamiento general (celdas combinadas)
    """
    # Cargar el archivo Excel
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)

    # Seleccionar la hoja
    if isinstance(hoja_origen, int):
        if len(wb.sheetnames) > hoja_origen:
            ws = wb.worksheets[hoja_origen]
        else:
            raise ValueError(f"No existe hoja con índice {hoja_origen}")
    else:
        if hoja_origen in wb.sheetnames:
            ws = wb[hoja_origen]
        else:
            raise ValueError(f"No existe hoja con nombre '{hoja_origen}'")

    # Convertir lista a conjunto para búsquedas más eficientes
    if ubicaciones_referencia and not isinstance(ubicaciones_referencia, set):
        ubicaciones_referencia = set(ubicaciones_referencia)

    ubicaciones = []
    celdas_procesadas = {}

    # 1. Procesar celdas combinadas
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds

        x = min_col - 1
        y = min_row - 1
        ancho = max_col - min_col + 1
        alto = max_row - min_row + 1

        valor = ws.cell(row=min_row, column=min_col).value
        ubicacion_valor, capacidad_valor = extraer_ubicacion_y_capacidad(valor)
        ubicacion_id = ubicacion_valor if ubicacion_valor else valor

        # Ignorar celdas específicas
        if valor is None or (
            isinstance(valor, str)
            and (
                valor.strip() == ""
                or valor.strip().startswith("P-")
                or valor.strip().startswith("PASILLO")
                or valor.strip().startswith("COLOSO")
            )
        ):
            continue

        if valor is None:
            col_letra = get_column_letter(min_col)
            valor = f"{col_letra}{min_row}"

        es_referencia = bool(
            ubicaciones_referencia and ubicacion_id in ubicaciones_referencia
        )

        ubicaciones.append(
            {
                "ubicacion_id": ubicacion_id,
                "x": x,
                "y": y,
                "ancho": ancho,
                "alto": alto,
                "valor_original": valor,
                "capacidad_max": capacidad_valor if capacidad_valor is not None else 0,
                "es_referencia": es_referencia,
            }
        )

        # Marcar celdas como procesadas
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                celdas_procesadas[(r, c)] = True

    # 2. Procesar celdas individuales
    max_row = ws.max_row
    max_col = ws.max_column

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            if (row, col) in celdas_procesadas:
                continue

            valor = ws.cell(row=row, column=col).value

            if valor is None or valor == "":
                continue
            if isinstance(valor, str) and (
                valor.strip() == ""
                or valor.strip().startswith("P-")
                or valor.strip().startswith("PASILLO")
                or valor.strip().startswith("COLOSO")
            ):
                continue

            if valor is not None:
                x = col - 1
                y = row - 1

                if isinstance(valor, str) and valor.strip():
                    ubicacion_id = valor
                else:
                    col_letra = get_column_letter(col)
                    ubicacion_id = f"{col_letra}{row}"

                ubicacion_valor, capacidad_valor = extraer_ubicacion_y_capacidad(valor)
                es_referencia = bool(
                    ubicaciones_referencia and ubicacion_id in ubicaciones_referencia
                )

                ubicaciones.append(
                    {
                        "ubicacion_id": ubicacion_id,
                        "x": x,
                        "y": y,
                        "ancho": 1,
                        "alto": 1,
                        "valor_original": valor,
                        "capacidad_max": (
                            capacidad_valor if capacidad_valor is not None else 0
                        ),
                        "es_referencia": es_referencia,
                    }
                )

                celdas_procesadas[(row, col)] = True

    return pd.DataFrame(ubicaciones)


def generar_layout_npr(uploaded_file, hoja_origen):
    """
    Genera layout desde Excel con procesamiento NPR
    """
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)

    # Seleccionar la hoja
    if isinstance(hoja_origen, int):
        if len(wb.sheetnames) > hoja_origen:
            ws = wb.worksheets[hoja_origen]
        else:
            raise ValueError(f"No existe hoja con índice {hoja_origen}")
    else:
        if hoja_origen in wb.sheetnames:
            ws = wb[hoja_origen]
        else:
            raise ValueError(f"No existe hoja con nombre '{hoja_origen}'")

    ubicaciones = []
    celdas_procesadas = {}

    # Procesar celdas combinadas (para marcar como procesadas)
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                celdas_procesadas[(r, c)] = True

    max_row = ws.max_row
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        pasillo = ws.cell(row=2, column=col).value  # S01, S02, etc.
        if pasillo is None:
            continue

        for row in range(1, max_row + 1):
            if (row, col) in celdas_procesadas:
                continue

            valor = ws.cell(row=row, column=col).value
            if valor is None or valor == "":
                continue

            if valor is not None:
                x = col - 1
                y = row - 1

                es_referencia = valor == "P"
                if str(valor).startswith("S"):
                    continue

                ubicacion = f"{pasillo}-{str(valor).zfill(2)}"

                ubicaciones.append(
                    {
                        "ubicacion_id": ubicacion,
                        "x": x,
                        "y": y,
                        "ancho": 1,
                        "alto": 1,
                        "valor_original": valor,
                        "capacidad_max": 1,
                        "es_referencia": es_referencia,
                    }
                )

                celdas_procesadas[(row, col)] = True

    return pd.DataFrame(ubicaciones)


def crear_svg_layout(df_layout, color_referencias="#FFD1A8"):
    """
    Crea un archivo SVG a partir del DataFrame del layout
    """
    if df_layout.empty:
        return "<svg></svg>"

    # Determinar dimensiones totales del layout
    max_x = df_layout["x"].max() + df_layout["ancho"].max()
    max_y = df_layout["y"].max() + df_layout["alto"].max()

    # Definir tamaño de celda y márgenes
    celda_ancho = 80
    celda_alto = 60
    margen = 20

    # Calcular dimensiones del SVG
    ancho_svg = (max_x + 1) * celda_ancho + 2 * margen
    alto_svg = (max_y + 1) * celda_alto + 2 * margen

    # Iniciar documento SVG
    svg = f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {ancho_svg} {alto_svg}" width="100%" height="100%">\n'
    svg += "  <style>\n"
    svg += "    .ubicacion-texto { font-family: Arial; font-size: 12px; text-anchor: middle; dominant-baseline: middle; fill: #000000; }\n"
    svg += f"    .ubicacion-referencia {{ fill: {color_referencias}; stroke: #000000; stroke-width: 2px; }}\n"
    svg += (
        "    .ubicacion-normal { fill: #ffffff; stroke: #000000; stroke-width: 1px; }\n"
    )
    svg += "    .ubicacion-normal:hover { fill: #ffcc80; cursor: pointer; }\n"
    svg += "  </style>\n"

    # Fondo
    svg += f'  <rect width="{ancho_svg}" height="{alto_svg}" fill="#f5f5f5" />\n'

    # Dibujar cada ubicación
    for _, ubicacion in df_layout.iterrows():
        x = margen + ubicacion["x"] * celda_ancho
        y = margen + ubicacion["y"] * celda_alto
        ancho = ubicacion["ancho"] * celda_ancho
        alto = ubicacion["alto"] * celda_alto
        ubicacion_id = ubicacion["ubicacion_id"]
        es_referencia = ubicacion.get("es_referencia", False)

        # Sanitizar el ID para uso en SVG
        ubicacion_id_safe = (
            str(ubicacion_id)
            .replace('"', "&quot;")
            .replace("'", "&apos;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace("&", "&amp;")
        )

        # Determinar la clase CSS según tipo de ubicación
        clase_css = "ubicacion-referencia" if es_referencia else "ubicacion-normal"

        # Crear el grupo para la ubicación
        svg += f'  <g id="{ubicacion_id_safe}" data-ubicacion="{ubicacion_id_safe}" data-capacidad="{ubicacion.get("capacidad_max", 1)}">\n'

        # Dibujar el rectángulo
        svg += f'    <rect class="{clase_css}" x="{x}" y="{y}" width="{ancho}" height="{alto}" rx="3" />\n'

        # Añadir el texto
        font_size = 10
        if len(str(ubicacion_id)) > 15 or (ancho < 80 or alto < 40):
            font_size = 8
        if len(str(ubicacion_id)) > 25 or (ancho < 50 or alto < 30):
            font_size = 6

        svg += f'    <text class="ubicacion-texto" x="{x + ancho/2}" y="{y + alto/2}" font-size="{font_size}px">{ubicacion_id}</text>\n'
        svg += "  </g>\n"

    # Finalizar SVG
    svg += "</svg>"

    return svg


def get_download_link(content, filename, mime_type):
    """Genera un enlace de descarga para el contenido"""
    b64 = base64.b64encode(content.encode()).decode()
    return f'<a href="data:{mime_type};base64,{b64}" download="{filename}">📥 Descargar {filename}</a>'


# Interfaz principal
def main():
    st.title("🗺️ Generador de Mapas SVG desde Excel")
    st.markdown("---")

    # Sidebar para configuración
    with st.sidebar:
        st.header("⚙️ Configuración")

        # Tipo de procesamiento
        tipo_procesamiento = st.selectbox(
            "Tipo de procesamiento:",
            ["General (Celdas combinadas)", "NPR (Pasillos numerados)"],
            help="Selecciona el tipo de procesamiento según tu archivo Excel",
        )

        # Configuración de colores
        st.subheader("🎨 Colores")
        color_referencias = st.color_picker("Color de referencias", "#FFD1A8")

        if tipo_procesamiento == "General (Celdas combinadas)":
            st.subheader("📍 Ubicaciones de Referencia")
            ubicaciones_ref_text = st.text_area(
                "Ubicaciones de referencia (una por línea):",
                value="INDUSTRIAL\nKIOSCO\nREEMPAQUE\nCARGA TRASERA\nRECEPCIÓN LATERAL",
                help="Ingresa las ubicaciones que deben marcarse como referencias",
            )
            ubicaciones_referencia = [
                u.strip() for u in ubicaciones_ref_text.split("\n") if u.strip()
            ]
        else:
            ubicaciones_referencia = ["P"]

    # Área principal
    col1, col2 = st.columns([1, 2])

    with col1:
        st.header("📤 Subir Archivo")

        # Upload de archivo
        uploaded_file = st.file_uploader(
            "Selecciona tu archivo Excel:",
            type=["xlsx", "xls"],
            help="Sube el archivo Excel con el layout de tu bodega",
        )

        if uploaded_file is not None:
            # Mostrar información del archivo
            st.success(f"✅ Archivo cargado: {uploaded_file.name}")

            # Leer hojas disponibles
            try:
                wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                hojas_disponibles = wb.sheetnames

                hoja_seleccionada = st.selectbox(
                    "Selecciona la hoja:",
                    hojas_disponibles,
                    help="Elige la hoja que contiene el layout",
                )

                # Botón para procesar
                if st.button("🚀 Generar Mapa SVG", type="primary"):
                    with st.spinner("Procesando archivo..."):
                        try:
                            # Resetear posición del archivo
                            uploaded_file.seek(0)

                            # Procesar según el tipo seleccionado
                            if tipo_procesamiento == "General (Celdas combinadas)":
                                df_layout = generar_layout_general(
                                    uploaded_file,
                                    ubicaciones_referencia,
                                    hoja_seleccionada,
                                )
                            else:
                                df_layout = generar_layout_npr(
                                    uploaded_file, hoja_seleccionada
                                )

                            # Guardar en session state
                            st.session_state.df_layout = df_layout
                            st.session_state.svg_content = crear_svg_layout(
                                df_layout, color_referencias
                            )

                            st.success("✅ ¡Mapa generado exitosamente!")

                            # Mostrar estadísticas
                            st.subheader("📊 Estadísticas")
                            st.metric("Total ubicaciones", len(df_layout))
                            referencias = df_layout[df_layout["es_referencia"] == True]
                            st.metric("Referencias", len(referencias))
                            normales = df_layout[df_layout["es_referencia"] == False]
                            st.metric("Ubicaciones normales", len(normales))

                        except Exception as e:
                            st.error(f"❌ Error al procesar el archivo: {str(e)}")

            except Exception as e:
                st.error(f"❌ Error al leer el archivo: {str(e)}")

    with col2:
        st.header("🗺️ Vista Previa del Mapa")

        if "svg_content" in st.session_state and "df_layout" in st.session_state:
            # Mostrar el SVG
            st.components.v1.html(
                f'<div style="width: 100%; height: 600px; overflow: auto; border: 1px solid #ddd; border-radius: 5px;">{st.session_state.svg_content}</div>',
                height=600,
            )

            # Opciones de descarga
            st.subheader("📥 Descargas")

            col_download1, col_download2, col_download3 = st.columns(3)

            with col_download1:
                # Descargar SVG
                svg_filename = f"mapa_layout_{uploaded_file.name.split('.')[0]}.svg"
                st.markdown(
                    get_download_link(
                        st.session_state.svg_content, svg_filename, "image/svg+xml"
                    ),
                    unsafe_allow_html=True,
                )

            with col_download2:
                # Descargar CSV
                csv_content = st.session_state.df_layout.to_csv(index=False)
                csv_filename = f"layout_data_{uploaded_file.name.split('.')[0]}.csv"
                st.markdown(
                    get_download_link(csv_content, csv_filename, "text/csv"),
                    unsafe_allow_html=True,
                )

            with col_download3:
                # Descargar Excel
                excel_buffer = io.BytesIO()
                st.session_state.df_layout.to_excel(
                    excel_buffer, index=False, sheet_name="Layout"
                )
                excel_content = base64.b64encode(excel_buffer.getvalue()).decode()
                excel_filename = f"layout_data_{uploaded_file.name.split('.')[0]}.xlsx"
                st.markdown(
                    f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{excel_content}" download="{excel_filename}">📥 Descargar {excel_filename}</a>',
                    unsafe_allow_html=True,
                )

            # Mostrar tabla de datos
            with st.expander("📋 Ver datos del layout"):
                st.dataframe(st.session_state.df_layout, use_container_width=True)

        else:
            st.info(
                "👆 Sube un archivo Excel y genera el mapa para ver la vista previa aquí."
            )

            # Mostrar ejemplo de cómo debe ser el archivo
            st.subheader("💡 Formato esperado del archivo")

            if tipo_procesamiento == "General (Celdas combinadas)":
                st.markdown(
                    """
                **Para procesamiento General:**
                - Celdas con nombres de ubicaciones
                - Celdas combinadas para áreas grandes
                - Formato: `UBICACION (CAPACIDAD)` opcional
                - Se ignoran celdas que empiecen con "P-", "PASILLO", "COLOSO"
                """
                )
            else:
                st.markdown(
                    """
                **Para procesamiento NPR:**
                - Fila 2: Códigos de pasillo (S01, S02, etc.)
                - Celdas con números de posición
                - Se genera formato: `PASILLO-POSICION`
                - "P" se marca como referencia
                """
                )


if __name__ == "__main__":
    main()
